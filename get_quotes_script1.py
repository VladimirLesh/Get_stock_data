from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
import os
import openpyxl
import platform

def notify_user(message):
    system_platform = platform.system()

    if system_platform == "Windows":
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, message, "Уведомление", 1)
    elif system_platform == "Darwin":  # macOS
        import subprocess
        subprocess.run(['osascript', '-e', f'display notification "{message}" with title "Уведомление"'])
    else:
        print(f"Уведомление: {message}")

def save_to_excel(dataframe, filename, sheetname):
    try:
        # Пытаемся загрузить существующий файл
        existing_data = pd.read_excel(filename, sheet_name=sheetname)

        # Добавляем новые данные
        updated_data = pd.concat([existing_data, dataframe], axis=1)

        # # Записываем обновленные данные в файл
        # with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        #     updated_data.to_excel(writer, sheet_name=sheetname, index=False)
    except FileNotFoundError:
        # Если файл не существует, создаем новый с указанным листом
        dataframe.to_excel(filename, sheet_name=sheetname, index=False)
    except ValueError:
        # Если лист не существует, создаем новый лист и добавляем данные
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            dataframe.to_excel(writer, sheet_name=sheetname, index=False)

def check_and_create_file(file_path):
    if not os.path.isfile(file_path):
        # Файл не существует, создаем его
        with open(file_path, 'w') as file:
            file.write("Привет, это новый файл!")

    print(f"Файл {file_path} проверен и готов к использованию.")

def checkRow(filename):
    df = pd.read_excel(filename, header=None)

    if not df.empty:
        last_row_index = df.iloc[:, 0].last_valid_index()
    else:
        last_row_index = None

    if last_row_index is not None:
        last_row = last_row_index + 1
    else:
        last_row = 0

    return last_row

def check_row(sheet):
    last_row_index = sheet.max_row
    return last_row_index


def create_index_sheet(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    if sheetname not in workbook.sheetnames:
        workbook.create_sheet(sheetname)
        workbook.save(filename)

def check_and_create_excel_file(filename):
    if not os.path.isfile(filename):
        df_empty = pd.DataFrame()
        df_empty.to_excel(filename, index=False)
        create_index_sheet(filename)

def check_and_save_file_excel(filename, names_companies, prices, atrs, sheetname):
    wb = None
    ws = None

    if os.path.isfile(filename):
        wb = openpyxl.load_workbook(filename)
        if sheetname in wb.sheetnames:
            ws = wb[sheetname]
        else:
            ws = wb.create_sheet(sheetname)
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(sheetname)

    last_row = check_row(ws)
    last_row = int(last_row)

    for name, price, atr in zip(names_companies, prices, atrs):
        last_row += 1
        ws[f'A{last_row}'] = name.text.strip()
        ws[f'B{last_row}'] = price.text.strip()
        ws[f'C{last_row}'] = atr.text.strip()
    wb.save(filename)

def check_row(sheet):
    last_row_index = sheet.max_row
    return last_row_index

def uploadInExcel(value, filename,  sheetname):
    webdriver_path = '/path/to/chromedriver'
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # Запуск браузера в фоновом режиме
    chrome_options.add_argument(f'--webdriver={webdriver_path}')  # Путь к драйверу
    driver = webdriver.Chrome(options=chrome_options)

    url = value
    driver.get(url)
    time.sleep(1)
    page_source = driver.page_source
    driver.quit()

    soup = BeautifulSoup(page_source, 'html.parser')

    namesCompanies = soup.find_all('span', class_='tv-screener__description')
    prices = soup.find_all('td',
                           class_='tv-data-table__cell tv-screener-table__cell tv-screener-table__cell--with-marker',
                           attrs={'title': '', 'data-field-key': 'close'})
    atrs = soup.find_all('td',
                         class_='tv-data-table__cell tv-screener-table__cell tv-screener-table__cell--down tv-screener-table__cell--with-marker',
                         attrs={'title': '', 'data-field-key': 'change'})

    data = []
    for name, price, atr in zip(namesCompanies, prices, atrs):
        data.append([name.text.strip(), price.text.strip(), atr.text.strip()])


    df = pd.DataFrame(data, columns=['Актив', 'Цена', 'ATR'])
    save_to_excel(df, filename, sheetname)
    # df.to_excel(filename,sheet_name=sheetname, index=False)
    print('Акции успешно импортированы! Ждем загрузку индексов...')

def uploadInExcelIndi(url_index, sheetname):
    webdriver_path = '/path/to/chromedriver'
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # Запуск браузера в фоновом режиме
    chrome_options.add_argument(f'--webdriver={webdriver_path}')  # Путь к драйверу
    driver = webdriver.Chrome(options=chrome_options)

    soup_array = []
    for index in url_index:
        url = index
        driver.get(url)
        time.sleep(1)
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        soup_array.append(soup)

    driver.quit()

    clear_excel_sheet(filename, sheetname)

    for soup in soup_array:
        namesCompanies = soup.find_all('h1', class_='apply-overflow-tooltip title-HFnhSVZy')
        prices = soup.find_all('span', class_='last-JWoJqCpY js-symbol-last')
        atrs = soup.find_all('span', class_='js-symbol-change-pt')
        check_and_save_file_excel(filename, namesCompanies, prices, atrs, sheetname)
        print(namesCompanies, 'Успех!')

def clear_excel_sheet(filename, sheetname):
    try:
        book = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        book = openpyxl.Workbook()

        # Выбираем лист или создаем новый
    sheet = book[sheetname] if sheetname in book.sheetnames else book.create_sheet(sheetname)

    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    book.save(filename)

filename = 'C:/Users/vladi/OneDrive/Рабочий стол/копия.xlsx'

url = 'https://ru.tradingview.com/screener/'
urlSP500 = 'https://ru.tradingview.com/symbols/SPX/?exchange=SP'
urlHangSeng = 'https://ru.tradingview.com/symbols/TVC-HSI/'
urlIMOEX = 'https://ru.tradingview.com/symbols/MOEX-IMOEX/'
urlRTSI = 'https://ru.tradingview.com/symbols/MOEX-RTSI/'
urlGC1 = 'https://ru.tradingview.com/symbols/COMEX-GC1!/'
urlGOLDRUBTOM = 'https://ru.tradingview.com/symbols/USDRUB_TOM/?exchange=MOEX'
urlUcloilBrent = 'https://ru.tradingview.com/symbols/UKOIL/?exchange=TVC'
urlES1 = 'https://ru.tradingview.com/symbols/CME_MINI-ES11!/?contract=ES101Z2023'
urlNG1 = 'https://ru.tradingview.com/symbols/MOEX-NG1!/'
urlEURUSD = 'https://ru.tradingview.com/symbols/EURUSD/?exchange=OANDA'
urlCNHUSD = 'https://ru.tradingview.com/symbols/CNHUSD/?exchange=FX_IDC'
urlFGBL1 = 'https://ru.tradingview.com/symbols/EUREX-FGBL1!/'
urlRGBI1 = 'https://ru.tradingview.com/symbols/MOEX-RGBI/'
urlETHUSD = 'https://ru.tradingview.com/symbols/ETHUSD/?exchange=BINANCE'
urlBTCUSD = 'https://ru.tradingview.com/symbols/BTCUSD/?exchange=BINANCE'
urlZN1 = 'https://ru.tradingview.com/symbols/CBOT-ZN1!/'
urlYNDX = 'https://ru.tradingview.com/symbols/MOEX-YNDX/'
urlOZON = 'https://ru.tradingview.com/symbols/MOEX-OZON/'

url_index = [urlSP500,urlHangSeng,urlIMOEX,urlRTSI,urlGC1,urlGOLDRUBTOM,urlUcloilBrent,
             urlES1,urlNG1,urlEURUSD,urlCNHUSD,urlFGBL1,urlRGBI1,urlETHUSD,urlBTCUSD,urlZN1]

urlStocks = [urlYNDX,urlOZON]

# create_index_sheet(filename, 'Акции')
# create_index_sheet(filename, 'Индексы')
uploadInExcel(url,filename, 'акции скринер')
uploadInExcelIndi(url_index, 'Индексы')
uploadInExcelIndi(urlStocks, 'акции')
notify_user('Данные успешно импортированы!')

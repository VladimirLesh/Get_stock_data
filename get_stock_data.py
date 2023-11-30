import pandas as pd
from io import StringIO
import requests
import openpyxl
import schedule
import time
import os


def get_stock_data():
    url = 'https://finviz.com/screener.ashx'
    response = requests.get(url)

    html_string = response.text
    df = pd.read_html(StringIO(html_string))[0]

    return df.iloc[:, :7]  # Выбираем все данные до столбца G включительно

def shift_and_save_to_excel(dataframe, filename='stock_data_lab7.xlsx'):
    # Проверяем существование файла
    file_exists = os.path.exists(filename)

    # Открываем файл Excel с помощью openpyxl
    if file_exists:
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()

    # Загружаем существующий лист 'StockData' или создаем новый
    try:
        sheet = workbook['StockData']
        if sheet.max_column > 6:
            # Переносим существующие данные вправо
            for col_num in range(sheet.max_column, 1, -1):
                for row_num in range(1, sheet.max_row + 1):
                    sheet.cell(row=row_num, column=col_num + 1, value=sheet.cell(row=row_num, column=col_num).value)

            # Записываем новые данные в столбец G
        for row_num in range(1, len(dataframe) + 1):
            sheet.cell(row=row_num, column=7, value=dataframe.iloc[row_num - 1, 6])
    except KeyError:
        sheet = workbook.create_sheet('StockData')
        for row_num in range(1, len(dataframe) + 1):
            for col_num in range(1, 8):  # Записываем данные от A до G
                sheet.cell(row=row_num, column=col_num, value=dataframe.iloc[row_num - 1, col_num - 1])


    # Сохраняем изменения
    workbook.save(filename)


def replace_dots_with_commas(filename='stock_data_lab7.xlsx'):
    file_exists = os.path.exists(filename)

    # Открываем файл Excel с помощью openpyxl
    if file_exists:
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()

    # Загружаем существующий лист 'StockData' или создаем новый
    try:
        sheet = workbook['StockData']
    except KeyError:
        sheet = workbook.create_sheet('StockData')

    # Проверяем наличие данных
    if sheet.max_column > 0 and sheet.max_row > 0:
        # Читаем данные из файла в DataFrame
        df = pd.DataFrame(sheet.iter_rows(values_only=True), columns=[cell.value for cell in sheet[1]])

        # Заменяем точки на запятые
        df = df.apply(lambda x: x.apply(lambda y: str(y).replace('.', ',')))

        # Записываем измененные данные в лист 'StockData'
        for row_num, values in enumerate(df.values, start=1):
            for col_num, value in enumerate(values, start=1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # Сохраняем изменения
        workbook.save(filename)
def job():
    print("Выполняем задачу...")
    stock_data = get_stock_data()
    shift_and_save_to_excel(stock_data)
    replace_dots_with_commas()
    print("Задача выполнена!")

job()

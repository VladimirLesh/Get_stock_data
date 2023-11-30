from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
import os
import openpyxl
import platform

filename = 'котировки.xlsx'
def notify_user(message):
    system_platform = platform.system()

    if system_platform == "Windows":
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, message, "Уведомление", 1)
    elif system_platform == "Darwin":  # macOS
        import subprocess
        subprocess.run(['osascript', '-e', f'display notification "{message}" with title "Уведомление"'])
    else:
        print(f"Уведомление: {message}")

def check_and_create_file(file_path):
    if not os.path.isfile(file_path):
        # Файл не существует, создаем его
        with open(file_path, 'w') as file:
            file.write("Привет, это новый файл!")

    print(f"Файл {file_path} проверен и готов к использованию.")

def checkRow(filename):
    df = pd.read_excel(filename, header=None)

    if not df.empty:
        last_row_index = df.iloc[:, 0].last_valid_index()
    else:
        last_row_index = None

    if last_row_index is not None:
        last_row = last_row_index + 1
    else:
        last_row = 0

    return last_row

def check_row(sheet):
    last_row_index = sheet.max_row
    return last_row_index

def create_index_sheet(filename):
    workbook = openpyxl.load_workbook(filename)
    if 'Индексы' not in workbook.sheetnames:
        workbook.create_sheet('Индексы')
        workbook.save(filename)

def check_and_create_excel_file(filename):
    if not os.path.isfile(filename):
        df_empty = pd.DataFrame()
        df_empty.to_excel(filename, index=False)
        create_index_sheet(filename)

def check_and_save_file_excel(filename, names_companies, prices, atrs):
    wb = None
    ws = None

    if os.path.isfile(filename):
        # Если файл существует, пытаемся прочитать вкладку "Индексы"
        wb = openpyxl.load_workbook(filename)
        if 'Индексы' in wb.sheetnames:
            ws = wb['Индексы']
        else:
            # Если вкладка "Индексы" не существует, создаем ее
            ws = wb.create_sheet('Индексы')
    else:
        # Если файл не существует, создаем новый с вкладкой "Индексы"
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Индексы')

    last_row = check_row(ws)

    last_row = int(last_row)

    for name, price, atr in zip(names_companies, prices, atrs):
        last_row += 1
        ws[f'A{last_row}'] = name.text.strip()
        ws[f'B{last_row}'] = price.text.strip()
        ws[f'C{last_row}'] = atr.text.strip()

    # Сохраняем изменения
    wb.save(filename)

def check_row(sheet):
    # Функция проверки последней занятой строки
    last_row_index = sheet.max_row
    return last_row_index

def uploadInExcel(value):
    webdriver_path = '/path/to/chromedriver'
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # Запуск браузера в фоновом режиме
    chrome_options.add_argument(f'--webdriver={webdriver_path}')  # Путь к драйверу
    driver = webdriver.Chrome(options=chrome_options)

    url = value
    driver.get(url)
    time.sleep(1)
    page_source = driver.page_source
    driver.quit()

    soup = BeautifulSoup(page_source, 'html.parser')

    namesCompanies = soup.find_all('span', class_='tv-screener__description')
    prices = soup.find_all('td',
                           class_='tv-data-table__cell tv-screener-table__cell tv-screener-table__cell--with-marker',
                           attrs={'title': '', 'data-field-key': 'close'})
    atrs = soup.find_all('td',
                         class_='tv-data-table__cell tv-screener-table__cell tv-screener-table__cell--down tv-screener-table__cell--with-marker',
                         attrs={'title': '', 'data-field-key': 'change'})

    data = []
    for name, price, atr in zip(namesCompanies, prices, atrs):
        data.append([name.text.strip(), price.text.strip(), atr.text.strip()])

    df = pd.DataFrame(data, columns=['Актив', 'Цена', 'ATR'])
    df.to_excel('котировки.xlsx',sheet_name='Акции', index=False)
    print('Успех!')

def uploadInExcelIndi(url_index):
    webdriver_path = '/path/to/chromedriver'
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # Запуск браузера в фоновом режиме
    chrome_options.add_argument(f'--webdriver={webdriver_path}')  # Путь к драйверу
    driver = webdriver.Chrome(options=chrome_options)

    soup_array = []
    for index in url_index:
        url = index
        driver.get(url)
        time.sleep(1)
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        soup_array.append(soup)

    driver.quit()


    for soup in soup_array:
        namesCompanies = soup.find_all('h1', class_='apply-overflow-tooltip title-HFnhSVZy')
        prices = soup.find_all('span', class_='last-JWoJqCpY js-symbol-last')
        atrs = soup.find_all('span', class_='js-symbol-change-pt')
        check_and_save_file_excel(filename, namesCompanies, prices, atrs)
        print(namesCompanies, 'Успех!')




url = 'https://ru.tradingview.com/screener/'
urlSP500 = 'https://ru.tradingview.com/symbols/SPX/?exchange=SP'
urlHangSeng = 'https://ru.tradingview.com/symbols/TVC-HSI/'
urlIMOEX = 'https://ru.tradingview.com/symbols/MOEX-IMOEX/'
urlRTSI = 'https://ru.tradingview.com/symbols/MOEX-RTSI/'
urlGC1 = 'https://ru.tradingview.com/symbols/COMEX-GC1!/'
urlGOLDRUBTOM = 'https://ru.tradingview.com/symbols/USDRUB_TOM/?exchange=MOEX'
urlUcloilBrent = 'https://ru.tradingview.com/symbols/UKOIL/?exchange=TVC'
urlES1 = 'https://ru.tradingview.com/symbols/CME_MINI-ES11!/?contract=ES101Z2023'
urlNG1 = 'https://ru.tradingview.com/symbols/MOEX-NG1!/'
urlEURUSD = 'https://ru.tradingview.com/symbols/EURUSD/?exchange=OANDA'
urlCNHUSD = 'https://ru.tradingview.com/symbols/CNHUSD/?exchange=FX_IDC'
urlFGBL1 = 'https://ru.tradingview.com/symbols/EUREX-FGBL1!/'
urlRGBI1 = 'https://ru.tradingview.com/symbols/MOEX-RGBI/'
urlETHUSD = 'https://ru.tradingview.com/symbols/ETHUSD/?exchange=BINANCE'
urlBTCUSD = 'https://ru.tradingview.com/symbols/BTCUSD/?exchange=BINANCE'
urlZN1 = 'https://ru.tradingview.com/symbols/CBOT-ZN1!/'

url_index = [urlSP500,urlHangSeng,urlIMOEX,urlRTSI,urlGC1,urlGOLDRUBTOM,urlUcloilBrent,
             urlES1,urlNG1,urlEURUSD,urlCNHUSD,urlFGBL1,urlRGBI1,urlETHUSD,urlBTCUSD,urlZN1]

# uploadInExcelIndi(urlSP500)
# uploadInExcelIndi(urlHangSeng)
# uploadInExcelIndi(urlIMOEX)
# uploadInExcelIndi(urlRTSI)
# uploadInExcelIndi(urlGC1)
# uploadInExcelIndi(urlGOLDRUBTOM)
# uploadInExcelIndi(urlUcloilBrent)
# uploadInExcelIndi(urlES1)
# uploadInExcelIndi(urlNG1)
# uploadInExcelIndi(urlEURUSD)
# uploadInExcelIndi(urlCNHUSD)
# uploadInExcelIndi(urlFGBL1)
# uploadInExcelIndi(urlRGBI1)
# uploadInExcelIndi(urlETHUSD)
# uploadInExcelIndi(urlBTCUSD)
# uploadInExcelIndi(urlZN1)


# uploadInExcel(url)
# create_index_sheet(filename)
# uploadInExcelIndi(url_index)
notify_user('Данные успешно импортированы!')